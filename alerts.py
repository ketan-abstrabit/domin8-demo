"""The alerting layer.

Every number these alerts need already existed — Movement, Inv. Level,
Inventory status, Ageing and the sell-through columns are computed by
stock_vs_sales.py and validated against the client's own workbook. What was
missing was anything that turned those columns into something that reaches a
person. That is all this module does.

Three ideas do the work:

DIGEST, NOT PER-SKU.  424 reorder emails is not a notification, it is a
spreadsheet with extra steps. One message per cycle, grouped by alert, top N
rows each, full detail in the attached workbook.

STATE.  The same 401 out-of-stock SKUs should not shout every fortnight. Each
alert carries a key, and each run classifies it NEW / ONGOING / RESOLVED
against the previous run. The digest leads with what changed.

VALUE, NOT COUNT.  Rows are ranked by rupees at risk, so a 400-unit dead line
outranks a 3-unit one. The brief's complaint is capital blocked in dead stock,
so capital is what the email leads with.

Usage:
    python alerts.py --report reports/output [--rules alert_rules.yaml]
                     [--state alert_state.json] [--email]
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

# Windows consoles default to cp1252, which cannot encode the rupee sign
# this pipeline prints. That killed a run on a developer machine while
# working fine in CI, where the console is UTF-8. Force UTF-8 and replace
# anything unprintable rather than raising: a report must not die over a
# currency symbol.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


SEVERITY_ORDER = {"high": 0, "medium": 1, "info": 2}
SEVERITY_COLOUR = {"high": "#b42318", "medium": "#b54708", "info": "#175cd3"}


# ---------------------------------------------------------------------------
# tolerant column access
#
# The workbook reproduces the client's headers verbatim, typos and double
# spaces included ("Last 15 days  Sales", "Overall \nSell-through"). Matching
# on a flattened name means a tidy-up on their side does not break alerting.
# ---------------------------------------------------------------------------

def _flat(s) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


class Cols:
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.map = {_flat(c): c for c in df.columns}

    def name(self, *candidates) -> str | None:
        for c in candidates:
            hit = self.map.get(_flat(c))
            if hit is not None:
                return hit
        return None

    def series(self, *candidates, default=0) -> pd.Series:
        hit = self.name(*candidates)
        if hit is None:
            return pd.Series([default] * len(self.df), index=self.df.index)
        return self.df[hit]

    def num(self, *candidates) -> pd.Series:
        return pd.to_numeric(self.series(*candidates), errors="coerce").fillna(0)

    def text(self, *candidates) -> pd.Series:
        return self.series(*candidates, default="").astype(str).str.strip()


# ---------------------------------------------------------------------------
# inputs
# ---------------------------------------------------------------------------

def load_rules(path: Path) -> dict:
    import yaml
    if not path.exists():
        raise SystemExit(f"rules file not found: {path}")
    rules = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if "alerts" not in rules:
        raise SystemExit(f"{path.name} has no 'alerts:' section")
    return rules


def newest_svs(report_dir: Path) -> Path:
    hits = sorted(report_dir.glob("Stock_vs_Sales*.xlsx"),
                  key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        raise SystemExit(f"no Stock_vs_Sales workbook in {report_dir}")
    return hits[0]


def sku_value(report_dir: Path, fallback_to_asp: bool = True) -> tuple[pd.Series, float]:
    """Rupees on hand per SKU, and the average selling price used as a fallback.

    Two sources of truth exist and neither is complete: the inventory facts
    carry a valuation for Uniware and Amazon but the store SOH files send units
    only. Where valuation is missing we price the units at what that SKU
    actually realised, and where it never sold, at the overall average. Which
    basis was used is reported per row so nobody mistakes an estimate for a
    ledger figure.
    """
    inv_path = report_dir / "fact_inventory.csv"
    sales_path = report_dir / "fact_sales.csv"

    value = pd.Series(dtype=float)
    if inv_path.exists():
        inv = pd.read_csv(inv_path)
        if {"master_sku", "value_on_hand"} <= set(inv.columns):
            value = (inv.groupby("master_sku")["value_on_hand"]
                        .sum(min_count=1).dropna())
            value = value[value > 0]

    asp_overall, asp_sku = 0.0, pd.Series(dtype=float)
    if fallback_to_asp and sales_path.exists():
        s = pd.read_csv(sales_path)
        if "flow" in s.columns:
            s = s[s["flow"] == "sell_out"]
        if {"master_sku", "net_value", "qty_sold"} <= set(s.columns):
            g = s.groupby("master_sku")[["net_value", "qty_sold"]].sum()
            g = g[g["qty_sold"] > 0]
            asp_sku = (g["net_value"] / g["qty_sold"]).replace(
                [float("inf"), float("-inf")], pd.NA).dropna()
            asp_sku = asp_sku[asp_sku > 0]
            tot_q, tot_v = g["qty_sold"].sum(), g["net_value"].sum()
            asp_overall = float(tot_v / tot_q) if tot_q else 0.0

    return value, (asp_sku, asp_overall)


# ---------------------------------------------------------------------------
# the frame every rule reads
# ---------------------------------------------------------------------------

def prepare(svs_path: Path, report_dir: Path, rules: dict) -> pd.DataFrame:
    raw = pd.read_excel(svs_path, sheet_name="sku wise")
    c = Cols(raw)

    d = pd.DataFrame(index=raw.index)
    d["sku"] = c.text("Sku Code")
    d["item"] = c.text("Item Name")
    d["category"] = c.text("Category")
    d["gender"] = c.text("Gender")
    d["ageing"] = c.text("Ageing")
    d["movement"] = c.text("Movement")
    d["overall_movement"] = c.text("Overall Movement")
    d["inv_level"] = c.text("Inv. Level")
    d["inv_status"] = c.text("Inventory status")
    d["reorder_status"] = c.text("Reorder Status")
    d["style_reorder"] = c.text("Style Re-order Status")

    d["stock"] = c.num("Total Inventory")
    d["monthly_ros"] = c.num("monthly ros")
    d["ros_month"] = c.num("ros/month")
    d["net_sale"] = c.num(*_period(raw, "Net sale"))
    d["returns"] = c.num(*_period(raw, "Returns"))
    d["overall_sale"] = c.num("overall net sale qty")
    d["overall_ret"] = c.num("Overall ret")
    d["sell_through"] = c.num(*_period(raw, "Sell through"))
    d["last15_sales"] = c.num("Last 15 days  Sales", "Last 15 days Sales")
    d["for_3_months"] = c.num("For 3 months")
    d["purchase_qty"] = c.num("Purchase. qty")

    d = d[d["sku"].astype(bool)].copy()

    # --- demand rate ---------------------------------------------------
    # `monthly ros` is derived from the last-15-day window and is legitimately
    # zero whenever the newest sales file predates that window — which is the
    # normal state a day after a cycle closes. Falling back to the 60-day rate
    # keeps every demand-side rule alive instead of silently matching nothing.
    d["demand_rate"] = d["monthly_ros"].where(d["monthly_ros"] > 0, d["ros_month"])
    daily = d["demand_rate"] / 30.0
    d["daily_rate"] = daily
    d["cover_days"] = (d["stock"] / daily).where(daily > 0)

    # --- return rate ---------------------------------------------------
    gross = d["overall_sale"] + d["overall_ret"]
    d["return_rate"] = (d["overall_ret"] / gross).where(gross > 0)

    # --- money ---------------------------------------------------------
    value, (asp_sku, asp_overall) = sku_value(
        report_dir, rules.get("valuation", {}).get("fallback_to_asp", True))
    on_hand = d["sku"].map(value)
    basis = pd.Series("reported", index=d.index).where(on_hand.notna(), "estimated")
    unit = d["sku"].map(asp_sku).fillna(asp_overall)
    est = d["stock"] * unit
    d["value_at_risk"] = on_hand.fillna(est).fillna(0.0).round(0)
    d["value_basis"] = basis.where(d["value_at_risk"] > 0, "none")

    # Two different questions need two different money columns. For dead,
    # ageing and slow stock the number that matters is capital sitting still —
    # value on hand. For out-of-stock, low stock and reorder the stock is by
    # definition nearly gone, so its value is near zero and ranking on it puts
    # the least urgent lines on top; what matters there is the revenue the
    # shortage costs per month. Each rule declares which one it ranks on.
    price = (d["value_at_risk"] / d["stock"]).where(d["stock"] > 0)
    d["unit_price"] = price.fillna(unit).fillna(asp_overall).fillna(0.0)
    d["revenue_at_risk"] = (d["demand_rate"] * d["unit_price"]).round(0)

    return d.reset_index(drop=True)


def _period(raw: pd.DataFrame, prefix: str) -> tuple[str, ...]:
    """The period columns are stamped with the window ("Net sale (Jul-Aug )").

    Return every header that starts with the prefix so Cols can match whichever
    month range this run produced.
    """
    hits = [c for c in raw.columns if _flat(c).startswith(_flat(prefix))]
    return tuple(hits) or (prefix,)


# ---------------------------------------------------------------------------
# the rules
#
# One function per alert. Each returns a boolean mask plus a per-row reason, so
# the digest can say *why* a line is on the list rather than just naming it.
# ---------------------------------------------------------------------------

def _fmt(n) -> str:
    if pd.isna(n):
        return "—"
    return f"{n:,.0f}" if abs(n) >= 10 else f"{n:,.1f}"


def rule_out_of_stock(d, cfg):
    m = (d["stock"] <= 0) & (d["demand_rate"] >= cfg.get("min_monthly_ros", 0.5))
    why = ("sells ~" + d["demand_rate"].map(_fmt) + "/month, stock is zero")
    return m, why


def rule_low_stock(d, cfg):
    """Two ways in, because the workbook and the maths disagree usefully.

    The merchandising workbook has its own 'low stock' bucket, which the team
    already trusts. Days-of-cover catches what that bucket misses: a SKU with a
    healthy-looking pile that is selling fast enough to be gone before the next
    PO lands. Either one is worth an alert, so take the union.
    """
    cover = cfg.get("cover_days", 21)
    by_cover = ((d["stock"] > 0)
                & (d["cover_days"].notna())
                & (d["cover_days"] < cover)
                & (d["demand_rate"] >= cfg.get("min_monthly_ros", 0.5)))
    by_status = (d["stock"] > 0) & d["inv_status"].str.lower().eq("low stock")
    m = by_cover | by_status
    why = pd.Series("flagged low stock in the workbook", index=d.index)
    why = why.where(~by_cover,
                    d["cover_days"].map(_fmt) + f" days of cover left (floor is {cover})")
    return m, why


def rule_reorder_level(d, cfg):
    m = d["inv_level"].str.upper().eq("REORDER")
    if cfg.get("respect_manual_override", True):
        done = d["reorder_status"].str.strip().str.lower()
        m &= ~done.isin({"ordered", "po raised", "done", "actioned", "no"})
    why = ("at reorder point — " + d["stock"].map(_fmt) + " on hand vs "
           + d["for_3_months"].map(_fmt) + " needed for 3 months")
    return m, why


def rule_best_sellers(d, cfg):
    m = (d["movement"].eq("Fast Movement")
         & (d["net_sale"] >= cfg.get("min_units_sold", 5)))
    why = (d["net_sale"].map(_fmt) + " sold this period, sell-through "
           + (d["sell_through"] * 100).map(_fmt) + "%")
    return m, why


def rule_high_returns(d, cfg):
    rate = cfg.get("return_rate", 0.15)
    m = ((d["return_rate"].notna())
         & (d["return_rate"] > rate)
         & (d["overall_sale"] + d["overall_ret"] >= cfg.get("min_units_sold", 10)))
    why = ((d["return_rate"] * 100).map(_fmt) + "% returned ("
           + d["overall_ret"].map(_fmt) + " of "
           + (d["overall_sale"] + d["overall_ret"]).map(_fmt) + ")")
    return m, why


def rule_slow_moving(d, cfg):
    m = (d["movement"].eq("Slow Movement")
         & (d["value_at_risk"] >= cfg.get("min_value", 5000)))
    why = ("sell-through " + (d["sell_through"] * 100).map(_fmt)
           + "% against its " + d["ageing"] + " age")
    return m, why


def rule_dead_stock(d, cfg):
    buckets = set(cfg.get("ageing_buckets", [">1 year", ">2 years"]))
    m = (d["movement"].eq("No Movement")
         & d["ageing"].isin(buckets)
         & (d["stock"] >= cfg.get("min_units", 5)))
    why = ("no sales, " + d["stock"].map(_fmt) + " units sitting "
           + d["ageing"].str.replace(">", "over ", regex=False))
    return m, why


def rule_stock_ageing(d, cfg):
    buckets = set(cfg.get("ageing_buckets", [">2 years"]))
    m = (d["ageing"].isin(buckets) & (d["stock"] >= cfg.get("min_units", 5)))
    why = (d["stock"].map(_fmt) + " units in the " + d["ageing"] + " bucket")
    return m, why


RULES = {
    "out_of_stock": rule_out_of_stock,
    "low_stock": rule_low_stock,
    "reorder_level": rule_reorder_level,
    "best_sellers": rule_best_sellers,
    "high_returns": rule_high_returns,
    "slow_moving": rule_slow_moving,
    "dead_stock": rule_dead_stock,
    "stock_ageing": rule_stock_ageing,
}


# ---------------------------------------------------------------------------
# evaluation
# ---------------------------------------------------------------------------

def evaluate(d: pd.DataFrame, rules: dict) -> pd.DataFrame:
    frames = []
    for key, cfg in (rules.get("alerts") or {}).items():
        cfg = cfg or {}
        if not cfg.get("enabled", True):
            continue
        fn = RULES.get(key)
        if fn is None:
            print(f"  ! unknown alert '{key}' in the rules file — skipped")
            continue
        mask, why = fn(d, cfg)
        mask = mask.fillna(False)
        if not mask.any():
            continue
        rows = d.loc[mask].copy()
        rows["alert"] = key
        rows["label"] = cfg.get("label", key.replace("_", " ").title())
        rows["severity"] = cfg.get("severity", "medium")
        rows["reason"] = why.loc[mask]

        rank_by = cfg.get("rank_by", "value_at_risk")
        if rank_by not in rows.columns:
            print(f"  ! '{key}' ranks on unknown column '{rank_by}' — "
                  "falling back to value at risk")
            rank_by = "value_at_risk"
        rows["priority"] = rows[rank_by]
        rows["money_col"] = rank_by
        rows["money_label"] = ("Monthly revenue at risk"
                               if rank_by == "revenue_at_risk"
                               else "Value at risk")
        frames.append(rows)

    if not frames:
        return pd.DataFrame(columns=list(d.columns) +
                            ["alert", "label", "severity", "reason", "key",
                             "priority", "money_col", "money_label"])

    out = pd.concat(frames, ignore_index=True)
    out["key"] = out["alert"] + "|" + out["sku"]
    out["sev_rank"] = out["severity"].map(SEVERITY_ORDER).fillna(9)
    return out.sort_values(["sev_rank", "alert", "priority"],
                           ascending=[True, True, False]).reset_index(drop=True)


def diff_state(current: pd.DataFrame, previous: dict, run_id: str) -> tuple[pd.DataFrame, list[dict], dict]:
    """Classify each alert NEW / ONGOING, and list what cleared since last run."""
    prev = previous.get("alerts", {}) if previous else {}
    now_keys = set(current["key"]) if len(current) else set()

    status, first_seen, cycles = [], [], []
    for k in current["key"] if len(current) else []:
        rec = prev.get(k)
        if rec:
            status.append("ONGOING")
            first_seen.append(rec.get("first_seen"))
            cycles.append(int(rec.get("cycles", 1)) + 1)
        else:
            status.append("NEW")
            first_seen.append(run_id)
            cycles.append(1)

    if len(current):
        current = current.copy()
        current["status"] = status
        current["first_seen"] = first_seen
        current["cycles"] = cycles

    resolved = []
    for k, rec in prev.items():
        if k not in now_keys:
            alert, _, sku = k.partition("|")
            resolved.append({
                "alert": alert, "sku": sku,
                "label": rec.get("label", alert),
                "item": rec.get("item", ""),
                "first_seen": rec.get("first_seen"),
                "cycles": rec.get("cycles", 1),
            })

    state = {"run_id": run_id, "alerts": {}}
    for _, r in current.iterrows() if len(current) else []:
        state["alerts"][r["key"]] = {
            "label": r["label"], "item": r["item"],
            "first_seen": r["first_seen"], "cycles": int(r["cycles"]),
            "priority": float(r["priority"]),
        }
    return current, resolved, state


# ---------------------------------------------------------------------------
# output
# ---------------------------------------------------------------------------

SHEET_COLS = [
    ("status", "Status"), ("label", "Alert"), ("severity", "Severity"),
    ("sku", "Sku Code"), ("item", "Item Name"), ("category", "Category"),
    ("stock", "Stock"), ("monthly_ros", "Monthly ROS"),
    ("cover_days", "Days cover"), ("sell_through", "Sell-through"),
    ("return_rate", "Return rate"), ("ageing", "Ageing"),
    ("movement", "Movement"), ("value_at_risk", "Value at risk"),
    ("revenue_at_risk", "Monthly revenue at risk"),
    ("value_basis", "Value basis"), ("reason", "Why"),
    ("cycles", "Cycles open"), ("first_seen", "First flagged"),
]


def write_workbook(current: pd.DataFrame, resolved: list[dict],
                   rules: dict, out_path: Path, run_id: str) -> Path:
    """One sheet per alert plus a summary. This is the full list; email is a digest."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def shaped(df: pd.DataFrame) -> pd.DataFrame:
        keep = [(k, h) for k, h in SHEET_COLS if k in df.columns]
        out = df[[k for k, _ in keep]].copy()
        out.columns = [h for _, h in keep]
        return out

    summary = []
    if len(current):
        for key, grp in current.groupby("alert", sort=False):
            cfg = (rules.get("alerts") or {}).get(key, {}) or {}
            summary.append({
                "Alert": grp["label"].iloc[0],
                "Severity": grp["severity"].iloc[0],
                "SKUs": len(grp),
                "New this run": int((grp["status"] == "NEW").sum()),
                "Units": int(grp["stock"].sum()),
                "Ranked on": grp["money_label"].iloc[0],
                "Money at stake": float(grp["priority"].sum()),
                "Why it matters": cfg.get("why", ""),
            })
    summary_df = pd.DataFrame(summary)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        if summary_df.empty:
            pd.DataFrame({"Alert": ["Nothing triggered this run."]}).to_excel(
                xw, sheet_name="Summary", index=False)
        else:
            summary_df.to_excel(xw, sheet_name="Summary", index=False)

        if len(current):
            new_only = current[current["status"] == "NEW"]
            if len(new_only):
                shaped(new_only).to_excel(xw, sheet_name="New this run", index=False)
            for key, grp in current.groupby("alert", sort=False):
                name = str(grp["label"].iloc[0])[:31]
                shaped(grp).to_excel(xw, sheet_name=name, index=False)

        if resolved:
            pd.DataFrame(resolved).to_excel(xw, sheet_name="Resolved", index=False)

        pd.DataFrame({
            "Setting": ["Run", "Rules version", "Alerts enabled"],
            "Value": [run_id, rules.get("version", 1),
                      ", ".join(k for k, v in (rules.get("alerts") or {}).items()
                                if (v or {}).get("enabled", True))],
        }).to_excel(xw, sheet_name="Run info", index=False)

        head_fill = PatternFill("solid", fgColor="1F2937")
        for ws in xw.book.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = head_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.freeze_panes = "A2"
            for i, col in enumerate(ws.iter_cols(), start=1):
                longest = max((len(str(c.value)) for c in col
                               if c.value is not None), default=8)
                ws.column_dimensions[get_column_letter(i)].width = min(
                    max(10, longest + 2), 52)
    return out_path


def digest_html(current: pd.DataFrame, resolved: list[dict], rules: dict,
                run_id: str, links: dict | None = None) -> str:
    top_n = (rules.get("digest") or {}).get("top_n", 15)
    e = html.escape

    new_n = int((current["status"] == "NEW").sum()) if len(current) else 0

    # Summed once per SKU, not once per alert. The same dead SKU is usually
    # flagged by dead_stock and stock_ageing both; adding its value twice would
    # inflate the only number anyone reads.
    def unique_money(side: str, col: str) -> float:
        if not len(current):
            return 0.0
        part = current[current["money_col"] == side]
        if not len(part):
            return 0.0
        return float(part.drop_duplicates("sku")[col].sum())

    capital = unique_money("value_at_risk", "value_at_risk")
    revenue = unique_money("revenue_at_risk", "revenue_at_risk")

    def rupees(v: float, per_month: bool = False) -> str:
        suffix = "/mo" if per_month else ""
        if v >= 100_000:
            return f"&#8377;{v/100_000:,.1f}L{suffix}"
        return f"&#8377;{v:,.0f}{suffix}"

    # Groups run high severity first, then whatever changed most, then by money.
    # Alphabetical order by alert key would be arbitrary, and the first block in
    # the email is the only one some people read.
    groups = []
    if len(current):
        for key, grp in current.groupby("alert", sort=False):
            groups.append((SEVERITY_ORDER.get(grp["severity"].iloc[0], 9),
                           -int((grp["status"] == "NEW").sum()),
                           -float(grp["priority"].sum()), key, grp))
        groups.sort(key=lambda t: t[:3])

    p = [f"""<!doctype html><html><body style="margin:0;background:#f3f4f6;
padding:24px;font:14px/1.5 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
color:#111827">
<div style="max-width:760px;margin:0 auto;background:#fff;border-radius:10px;
overflow:hidden;border:1px solid #e5e7eb">
<div style="background:#111827;color:#fff;padding:20px 24px">
  <div style="font-size:18px;font-weight:600">DOMIN8 — stock alerts</div>
  <div style="opacity:.75;font-size:13px;margin-top:4px">{e(run_id)}</div>
</div>
<div style="padding:20px 24px">
<table role="presentation" style="width:100%;border-collapse:collapse;margin-bottom:20px">
<tr>
  <td style="padding:12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px">
    <div style="font-size:22px;font-weight:600">{new_n:,}</div>
    <div style="color:#6b7280;font-size:12px">new since last run</div></td>
  <td style="width:10px"></td>
  <td style="padding:12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px">
    <div style="font-size:22px;font-weight:600">{len(current):,}</div>
    <div style="color:#6b7280;font-size:12px">open alerts</div></td>
  <td style="width:10px"></td>
  <td style="padding:12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px">
    <div style="font-size:22px;font-weight:600">{rupees(capital)}</div>
    <div style="color:#6b7280;font-size:12px">capital stuck</div></td>
  <td style="width:10px"></td>
  <td style="padding:12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px">
    <div style="font-size:22px;font-weight:600">{rupees(revenue)}</div>
    <div style="color:#6b7280;font-size:12px">sales at risk / month</div></td>
  <td style="width:10px"></td>
  <td style="padding:12px;background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px">
    <div style="font-size:22px;font-weight:600">{len(resolved):,}</div>
    <div style="color:#6b7280;font-size:12px">cleared</div></td>
</tr></table>"""]

    if not len(current):
        p.append('<p style="color:#6b7280">Nothing triggered this run.</p>')

    for _, _, _, key, grp in groups:
        cfg = (rules.get("alerts") or {}).get(key, {}) or {}
        sev = grp["severity"].iloc[0]
        colour = SEVERITY_COLOUR.get(sev, "#374151")
        n_new = int((grp["status"] == "NEW").sum())
        per_mo = grp["money_col"].iloc[0] == "revenue_at_risk"
        val = rupees(float(grp["priority"].sum()), per_mo)
        money_head = "&#8377; / month" if per_mo else "&#8377;"
        noun = "SKU" if len(grp) == 1 else "SKUs"
        p.append(f"""
<div style="margin:22px 0 8px;padding-left:10px;border-left:3px solid {colour}">
  <div style="font-weight:600;font-size:15px">{e(str(grp['label'].iloc[0]))}
    <span style="color:#6b7280;font-weight:400">— {len(grp):,} {noun}</span>
    {f'<span style="color:{colour};font-weight:600"> · {n_new} new</span>' if n_new else ''}
    <span style="color:#6b7280;font-weight:400"> · {val}</span>
  </div>
  <div style="color:#6b7280;font-size:12.5px;margin-top:2px">{e(cfg.get('why',''))}</div>
</div>
<table style="width:100%;border-collapse:collapse;font-size:12.5px">
<tr style="text-align:left;color:#6b7280">
  <th style="padding:6px 8px;border-bottom:1px solid #e5e7eb">SKU</th>
  <th style="padding:6px 8px;border-bottom:1px solid #e5e7eb">Item</th>
  <th style="padding:6px 8px;border-bottom:1px solid #e5e7eb;text-align:right">Stock</th>
  <th style="padding:6px 8px;border-bottom:1px solid #e5e7eb;text-align:right">{money_head}</th>
  <th style="padding:6px 8px;border-bottom:1px solid #e5e7eb">Why</th></tr>""")
        for _, r in grp.head(top_n).iterrows():
            tag = ('<span style="color:#b42318;font-weight:600">NEW </span>'
                   if r["status"] == "NEW" else "")
            p.append(f"""<tr>
  <td style="padding:6px 8px;border-bottom:1px solid #f3f4f6;white-space:nowrap">{tag}{e(str(r['sku']))}</td>
  <td style="padding:6px 8px;border-bottom:1px solid #f3f4f6">{e(str(r['item'])[:44])}</td>
  <td style="padding:6px 8px;border-bottom:1px solid #f3f4f6;text-align:right">{r['stock']:,.0f}</td>
  <td style="padding:6px 8px;border-bottom:1px solid #f3f4f6;text-align:right">{r['priority']:,.0f}</td>
  <td style="padding:6px 8px;border-bottom:1px solid #f3f4f6;color:#6b7280">{e(str(r['reason']))}</td></tr>""")
        p.append("</table>")
        if len(grp) > top_n:
            p.append(f'<div style="color:#6b7280;font-size:12px;padding:6px 8px">'
                     f'+ {len(grp)-top_n:,} more in the attached workbook</div>')

    if resolved:
        names = ", ".join(f"{r['sku']}" for r in resolved[:12])
        more = f" and {len(resolved)-12} more" if len(resolved) > 12 else ""
        p.append(f"""<div style="margin-top:26px;padding:12px;background:#f0fdf4;
border:1px solid #bbf7d0;border-radius:6px;font-size:12.5px">
<b>Cleared since the last run ({len(resolved)}):</b> {e(names)}{more}</div>""")

    if links:
        rows = "".join(
            f'<div style="padding:2px 0"><a href="{e(u)}" style="color:#175cd3">{e(n)}</a></div>'
            for n, u in sorted(links.items()))
        p.append(f'<div style="margin-top:22px;font-size:12.5px">'
                 f'<b>This cycle&rsquo;s reports</b>{rows}</div>')

    p.append("""<div style="margin-top:26px;padding-top:14px;border-top:1px solid #e5e7eb;
color:#9ca3af;font-size:11.5px">
Thresholds live in <code>alert_rules.yaml</code>. Value at risk uses reported
inventory valuation where the source provides one and average realised price
where it does not — the workbook says which, per row.
</div></div></div></body></html>""")
    return "".join(p)


# ---------------------------------------------------------------------------
# email
# ---------------------------------------------------------------------------

def env(name: str, default: str = "") -> str:
    """os.environ.get, but an empty value counts as unset.

    GitHub Actions maps a secret that does not exist to the empty string rather
    than leaving the variable out, so `os.environ.get("SMTP_PORT", "587")`
    returns "" and never the default. Every optional setting has to go through
    this or the defaults silently do not apply.
    """
    import os
    return (os.environ.get(name) or "").strip() or default


def send_email(subject: str, body_html: str, recipients: list[str],
               attachment: Path | None = None, log=print) -> bool:
    """SMTP, so it works with a Workspace mailbox or a Gmail app password."""
    import smtplib
    from email.message import EmailMessage

    user = env("SMTP_USER")
    pw = env("SMTP_PASS")

    # Check before parsing anything. Reading the port first meant an unset
    # SMTP_PORT blew up before this guard could report the real situation.
    if not (user and pw and recipients):
        missing = [n for n, v in (("SMTP_USER", user), ("SMTP_PASS", pw),
                                  ("recipients", recipients)) if not v]
        log(f"  email not sent — not configured ({', '.join(missing)})")
        return False

    host = env("SMTP_HOST", "smtp.gmail.com")
    raw_port = env("SMTP_PORT", "587")
    try:
        port = int(raw_port)
    except ValueError:
        log(f"  SMTP_PORT is not a number ({raw_port!r}) — using 587")
        port = 587
    sender = env("SMTP_FROM", user)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.set_content("This report is formatted as HTML. "
                    "The full list is in the attached workbook.")
    msg.add_alternative(body_html, subtype="html")

    if attachment and attachment.exists():
        msg.add_attachment(
            attachment.read_bytes(), maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment.name)

    with smtplib.SMTP(host, port, timeout=60) as s:
        s.starttls()
        s.login(user, pw)
        s.send_message(msg)
    log(f"  email sent to {len(recipients)} recipient(s)")
    return True


def recipients_from(rules: dict) -> list[str]:
    listed = list((rules.get("digest") or {}).get("to") or [])
    listed += [a.strip() for a in re.split(r"[,;\s]+", env("ALERT_RECIPIENTS"))
               if a.strip()]
    seen, out = set(), []
    for a in listed:
        if a.lower() not in seen:
            seen.add(a.lower())
            out.append(a)
    return out


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------

def run(report_dir: Path, rules_path: Path, state_path: Path,
        run_id: str | None = None, email: bool = False,
        links: dict | None = None, log=print) -> dict:
    rules = load_rules(rules_path)
    svs = newest_svs(report_dir)
    run_id = run_id or datetime.now().strftime("%Y-%m-%d %H:%M")

    log(f"  reading {svs.name}")
    d = prepare(svs, report_dir, rules)
    current = evaluate(d, rules)

    previous = {}
    if state_path.exists():
        try:
            previous = json.loads(state_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            log("  ! alert_state.json unreadable — treating every alert as new")

    current, resolved, state = diff_state(current, previous, run_id)
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(state, indent=2), encoding="utf-8")

    wb = write_workbook(current, resolved, rules,
                        report_dir / "Alerts.xlsx", run_id)
    body = digest_html(current, resolved, rules, run_id, links)
    (report_dir / "alert_digest.html").write_text(body, encoding="utf-8")

    new_n = int((current["status"] == "NEW").sum()) if len(current) else 0
    log(f"  {len(current):,} open alerts, {new_n:,} new, {len(resolved):,} cleared")
    if len(current):
        for key, grp in current.groupby("alert", sort=False):
            n_new = int((grp["status"] == "NEW").sum())
            tag = "/mo" if grp["money_col"].iloc[0] == "revenue_at_risk" else "   "
            log(f"      {grp['label'].iloc[0]:<24} {len(grp):>5,}"
                f"  (+{n_new} new)  Rs {grp['priority'].sum():>12,.0f}{tag}")

    sent = False
    if email:
        dcfg = rules.get("digest") or {}
        quiet = not dcfg.get("send_when_no_change", False)
        if quiet and new_n == 0 and not resolved:
            log("  nothing changed — digest not sent (send_when_no_change: false)")
        else:
            subject = dcfg.get("subject", "DOMIN8 stock alerts — {date}").format(
                date=run_id, new=new_n, open=len(current))
            # Email is the last and least important thing this does, and it
            # depends on a third party. It must never take the cycle down with
            # it: the reports are already built and still need publishing.
            try:
                sent = send_email(subject, body, recipients_from(rules), wb, log=log)
            except Exception as exc:                            # noqa: BLE001
                log(f"  ! digest not sent: {type(exc).__name__}: {exc}")
                log("    the reports are unaffected and will still be published")

    capital = 0.0
    if len(current):
        stock_side = current[current["money_col"] == "value_at_risk"]
        capital = float(stock_side.drop_duplicates("sku")["value_at_risk"].sum())

    return {"open": len(current), "new": new_n, "resolved": len(resolved),
            "capital_at_risk": capital,
            "workbook": wb, "digest": report_dir / "alert_digest.html",
            "emailed": sent}


def main():
    ap = argparse.ArgumentParser(description="DOMIN8 alert engine")
    here = Path(__file__).resolve().parent
    ap.add_argument("--report", type=Path, default=here / "reports" / "output")
    ap.add_argument("--rules", type=Path, default=here / "alert_rules.yaml")
    ap.add_argument("--state", type=Path,
                    default=here / "reports" / "_state" / "alert_state.json")
    ap.add_argument("--run-id")
    ap.add_argument("--email", action="store_true")
    a = ap.parse_args()

    print("\nAlerts")
    r = run(a.report, a.rules, a.state, a.run_id, a.email)
    print(f"\n  wrote {r['workbook'].name} and {r['digest'].name}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
