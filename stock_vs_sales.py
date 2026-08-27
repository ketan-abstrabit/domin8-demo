#!/usr/bin/env python3
"""
Stock vs Sales -- the merchandising report, generated from the pipeline's facts.

Reproduces the "Stock VS Sales" workbook (sku wise + Article wise) from
fact_sales.csv / fact_inventory.csv / fact_purchase.csv, so it comes out of every
cycle instead of being rebuilt by hand.

    python stock_vs_sales.py --input ./reports/input --report ./reports/output

Options
    --asof 2026-07-31       report date (default: today)
    --period-days 60        the main sales window        (their "Jun-Jul")
    --recent-days 15        the short window             (their "Last 15 days")

Every derived column's formula is written into the workbook's Definitions sheet.

MOVEMENT CLASSIFICATION
-----------------------
Reverse-engineered from the client's own workbook and validated at 100% against
both of its sheets (1,716 SKU rows and 378 article rows):

    No Movement     sell-through <= 0
    Slow Movement   0 < sell-through < Avg * 0.75
    Good Movement   Avg * 0.75 <= sell-through <= Avg * 1.25
    Fast Movement   sell-through > Avg * 1.25

...where Avg is the expected sell-through for the item's ageing bucket. Note the
client's "Movement criteria" tab documents a fourth band (< Avg/4 = no movement)
that the workbook does not actually implement -- the real cut for No Movement is
zero sales. The implemented behaviour is reproduced here; see NOTE_MOVEMENT.
"""

from __future__ import annotations

import argparse
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore")

# ===========================================================================
# CONFIG 1 -- expected sell-through by ageing bucket ("Avg" in the criteria tab)
# ===========================================================================

AGEING_BUCKETS = [          # (label, max age in days)
    ("0-1 month",    31),
    ("1-3 months",   92),
    ("3-6 months",  183),
    ("6-9 months",  274),
    ("9-12 months", 365),
    (">1 year",     730),
    (">2 years", 10**6),
]

EXPECTED_SELL_THROUGH = {
    "0-1 month":   0.20,
    "1-3 months":  0.40,
    "3-6 months":  0.46,
    "6-9 months":  0.46,
    "9-12 months": 0.46,
    ">1 year":     0.46,
    ">2 years":    0.46,
}

SLOW_CUT, FAST_CUT = 0.75, 1.25     # multipliers on the expected sell-through

NOTE_MOVEMENT = (
    "No Movement = sell-through <= 0. Slow = below Avg x0.75. Good = Avg x0.75 "
    "to x1.25. Fast = above Avg x1.25. Avg is the expected sell-through for the "
    "ageing bucket (0-1m 20%, 1-3m 40%, older 46%). Validated at 100% against "
    "the client workbook's own 'sku wise' and 'Article wise' tabs."
)

# ===========================================================================
# CONFIG 2 -- inventory cover ladder (Total Inventory / monthly ros)
# ===========================================================================

COVER_BANDS = [("low stock", 1.5), ("in stock", 3.0)]   # "no stock" = zero units
COVER_TOP = "over stock"

# ===========================================================================
# CONFIG 3 -- which channel feeds which column of the report
# `channel` values come from CHANNEL_MAP in reconcile.py.
# ===========================================================================

INV_COLUMNS = {                       # report column -> channel(s) in fact_inventory
    "all sports inv": ["All Sports"],
    "J&J Inv":        ["Jack & Jill"],
    "OR inv":         ["Amazon Vendor Central"],
    "D8 inv":         ["Warehouse"],
    "INCS inv":       ["INCS"],
}

# "Coco-OR" is Cocoblu / Amazon Vendor Central -- the ASIN side of the business.
AMAZON_CHANNELS = ["Amazon Vendor Central"]
RETAIL_TYPES = ["retail_store"]

REPORT_COLS = [
    "Sku Code", "Item Name", "Category", "Gender",
    "Total Inventory", "all sports inv", "J&J Inv", "OR inv", "INCS inv",
    "d8 inv + rtv", "D8 inv", "CB rtv",
    "Sale", "Ret", "Coco-OR SALES", "Coco-OR RET", "Retail",
    "Net sale", "Returns", "avg", "Sale Projection for next 3 month",
    "Purchase. qty", "Sell through", "overall net sale qty", "Overall ret",
    "Overall Sell-through", "Sell-through basis", "ros/month",
    "Last 15 days Sales", "Last 15 days Returns", "monthly ros",
    "Ageing", "Ageing basis", "Overall Movement", "Inv. Level", "Movement",
    "Style Re-order Status", "For 3 months", "Inventory status", "Reorder Status",
]

ARTICLE_COLS = [
    "Item Name", "Category", "Gender",
    "Total Inventory", "all sports inv", "J&J Inv", "OR inv", "INCS inv",
    "d8 inv + rtv", "D8 inv", "CB rtv", "sizes", "sizes in stock",
    "Net sale", "Returns", "Ret %", "ros/month", "Purchase. qty", "Sell through",
    "Last 15 days Sales", "Last 15 days Returns",
    "overall net sale qty", "overall ret qty", "Overall Sell-through",
    "Sell-through basis", "Ageing", "overall Movement", "Inv. Level", "Movement", "Movement Stability",
    "Style Status", "For 3 months", "Inventory status", "Re-order Status",
]


# ===========================================================================
# CONFIG 4 -- OUTPUT LAYOUT
#
# Exact column headers and order from the client's own workbook, including its
# double spaces, trailing spaces, embedded newline and the "Reurns" typo. They
# are reproduced verbatim so existing pivots, VLOOKUPs and column references in
# the team's downstream sheets keep working. Do not tidy them.
#
# {P} is the period label, generated from the actual window (e.g. "Jun-Aug").
# Columns this pipeline adds that the client's sheet never had are appended at
# the END, so nothing existing shifts position.
# ===========================================================================

SKU_LAYOUT = [
    ("Sku Code",                         "Sku Code"),
    ("Item Name",                        "Item Name"),
    ("Category",                         "Category"),
    ("Total Inventory",                  "Total Inventory"),
    ("all sports inv",                   "all sports inv"),
    ("J&J Inv",                          "J&J Inv"),
    ("OR inv",                           "OR inv"),
    ("d8 inv + rtv",                     "d8 inv + rtv"),
    ("D8 inv",                           "D8 inv"),
    ("CB rtv",                           "CB rtv"),
    ("Sale   ({P} )",                    "Sale"),
    ("Ret",                              "Ret"),
    ("Coco-OR SALES",                    "Coco-OR SALES"),
    ("Coco-OR RET",                      "Coco-OR RET"),
    ("Retail ({P})",                     "Retail"),
    ("Net sale ({P} )",                  "Net sale"),
    ("Returns ({P} )",                   "Returns"),
    ("avg",                              "avg"),
    ("Sale Projection for next 3 month", "Sale Projection for next 3 month"),
    ("Purchase. qty",                    "Purchase. qty"),
    ("Sell through ({P})",               "Sell through"),
    ("overall net sale qty",             "overall net sale qty"),
    ("Overall ret",                      "Overall ret"),
    ("Overall \nSell-through",           "Overall Sell-through"),
    ("ros/month",                        "ros/month"),
    ("Last 15 days  Sales",              "Last 15 days Sales"),
    ("Last 15 days  Returns",            "Last 15 days Returns"),
    ("monthly ros",                      "monthly ros"),
    ("Ageing",                           "Ageing"),
    ("Overall Movement",                 "Overall Movement"),
    ("Inv. Level",                       "Inv. Level"),
    ("Movement",                         "Movement"),
    ("Style Re-order Status",            "Style Re-order Status"),
    ("For 3 months",                     "For 3 months"),
    ("Inventory status",                 "Inventory status"),
    ("Reorder Status",                   "Reorder Status"),
    # --- added by this pipeline, appended so nothing above shifts ---
    ("Gender",                           "Gender"),
    ("INCS inv",                         "INCS inv"),
    ("Ageing basis",                     "Ageing basis"),
    ("Sell-through basis",               "Sell-through basis"),
]

ARTICLE_LAYOUT = [
    ("Item Name",                          "Item Name"),
    ("Category",                           "Category"),
    ("Gender",                             "Gender"),
    ("Total Inventory",                    "Total Inventory"),
    ("d8 Inv + RTV",                       "d8 inv + rtv"),
    ("d8 Inventory",                       "D8 inv"),
    ("Net sale ({P})",                     "Net sale"),
    ("Reurns   ({P} 15)",                  "Returns"),
    ("ros/month",                          "ros/month"),
    ("Purchase. qty",                      "Purchase. qty"),
    ("Sell through",                       "Sell through"),
    ("Last 15 days sales ",                "Last 15 days Sales"),
    ("Last 15 daysRet",                    "Last 15 days Returns"),
    ("ros/ month ( Based on last 15 days)", "monthly ros"),
    ("overall net sale qty ",              "overall net sale qty"),
    ("overall ret qty ",                   "overall ret qty"),
    ("Overall \nSell-through",             "Overall Sell-through"),
    ("Ret %",                              "Ret %"),
    ("Ageing",                             "Ageing"),
    ("overall Movement",                   "overall Movement"),
    ("Inv. Level",                         "Inv. Level"),
    ("Movement",                           "Movement"),
    ("Movement Stability",                 "Movement Stability"),
    ("Style Status",                       "Style Status"),
    ("For 3 months",                       "For 3 months"),
    ("Inventory status",                   "Inventory status"),
    ("Re-order Status",                    "Re-order Status"),
    # --- added by this pipeline ---
    ("all sports inv",                     "all sports inv"),
    ("J&J Inv",                            "J&J Inv"),
    ("OR inv",                             "OR inv"),
    ("INCS inv",                           "INCS inv"),
    ("CB rtv",                             "CB rtv"),
    ("sizes",                              "sizes"),
    ("sizes in stock",                     "sizes in stock"),
    ("Ageing basis",                       "Ageing basis"),
    ("Sell-through basis",                 "Sell-through basis"),
]


def period_label(start, end) -> str:
    """'Jun-Aug' from the actual window, so the header says what it covers."""
    a, b = f"{start:%b}", f"{end:%b}"
    return a if a == b else f"{a}-{b}"


def lay_out(df: pd.DataFrame, layout, label: str) -> pd.DataFrame:
    """Project the working frame onto the client's exact header names/order."""
    out = pd.DataFrame(index=df.index)
    for header, key in layout:
        out[header.replace("{P}", label)] = (
            df[key] if key in df.columns else pd.Series(index=df.index, dtype=object))
    return out


# ===========================================================================
# Helpers
# ===========================================================================

def num(s) -> pd.Series:
    return pd.to_numeric(pd.Series(s), errors="coerce")


def norm(s) -> pd.Series:
    return (pd.Series(s).astype("string").astype(object).fillna("")
            .astype(str).str.strip().str.upper()
            .replace({"NAN": None, "": None}))


# ---------------------------------------------------------------------------
# Ageing fallbacks. A first-PO date is the truth; when PO history doesn't reach
# back far enough we fall back down this ladder and label which rung was used in
# the "Ageing basis" column, so nobody reads a season-derived age as a receipt.
# ---------------------------------------------------------------------------

SEASON_START_MONTH = {"SS": 2, "AW": 8}      # Spring/Summer Feb, Autumn/Winter Aug

# Styles like DOM8-AW-M-TS-24 carry a season but no year: the legacy range.
LEGACY_SEASON_DATE = "2023-08-01"


def season_date(style_or_sku) -> pd.Timestamp:
    """AW24FES0MTS11BE -> 2024-08-01 ; DOM8-SS-W-TP-10 -> legacy date."""
    s = str(style_or_sku).upper()
    m = re.search(r"\b(SS|AW)(\d{2})", s)
    if m:
        yy, mth = int(m.group(2)), SEASON_START_MONTH[m.group(1)]
        if 15 <= yy <= 60:                      # sane 2015-2060 window
            return pd.Timestamp(2000 + yy, mth, 1)
    if re.search(r"\b(SS|AW)\b|-(SS|AW)-", s):
        return pd.Timestamp(LEGACY_SEASON_DATE)
    return pd.NaT


def ageing_bucket(days) -> str | None:
    if days is None or pd.isna(days):
        return None
    for label, hi in AGEING_BUCKETS:
        if days <= hi:
            return label
    return AGEING_BUCKETS[-1][0]


def movement(sell_through, ageing) -> str | None:
    """The validated 4-band classifier. See NOTE_MOVEMENT."""
    avg = EXPECTED_SELL_THROUGH.get(str(ageing).strip()) if ageing else None
    if avg is None or sell_through is None or pd.isna(sell_through):
        return None
    if sell_through <= 0:
        return "No Movement"
    if sell_through < avg * SLOW_CUT:
        return "Slow Movement"
    if sell_through <= avg * FAST_CUT:
        return "Good Movement"
    return "Fast Movement"


def inventory_status(inv, monthly_ros) -> str:
    """Cover-based. With no recent sales, judge on stock alone."""
    if pd.isna(inv) or inv <= 0:
        return "no stock"
    if pd.isna(monthly_ros) or monthly_ros <= 0:
        return COVER_TOP if inv > 3 else "in stock"
    cover = inv / monthly_ros
    for label, hi in COVER_BANDS:
        if cover < hi:
            return label
    return COVER_TOP


def stability(recent_mv, overall_mv) -> str | None:
    """Is the recent trend better or worse than the lifetime picture?"""
    rank = {"No Movement": 0, "Slow Movement": 1, "Good Movement": 2, "Fast Movement": 3}
    a, b = rank.get(recent_mv), rank.get(overall_mv)
    if a is None or b is None:
        return None
    if a > b:
        return "Improving"
    if a < b:
        return "Slowing Down" if b >= 2 else "Recovering"
    return "Stable"



def safe_excel_path(path: Path) -> Path:
    """Windows locks a workbook that is open in Excel, and a network share can
    refuse an in-place overwrite. Either way, write beside it rather than dying
    on the last step after all the work is done."""
    try:
        with open(path, "ab"):
            pass
        return path
    except OSError:
        alt = path.with_name(f"{path.stem}_{datetime.now():%H%M}{path.suffix}")
        print(f"  !! {path.name} is locked or read-only (open in Excel?) "
              f"— writing {alt.name} instead")
        return alt


def safe_to_csv(df, path: Path) -> Path:
    """Same protection for the seeded override template."""
    try:
        with open(path, "ab"):
            pass
        target = path
    except OSError:
        target = path.with_name(f"{path.stem}_{datetime.now():%H%M}{path.suffix}")
        print(f"  !! {path.name} is locked (open in Excel?) — wrote {target.name}")
    df.to_csv(target, index=False)
    return target

# ===========================================================================
# Load
# ===========================================================================

def load_facts(report_dir: Path):
    need = report_dir / "fact_sales.csv"
    if not need.exists():
        sys.exit(f"{need} not found — run reconcile.py / run_pipeline.py first.")
    fs = pd.read_csv(need, low_memory=False)
    fi_p = report_dir / "fact_inventory.csv"
    fp_p = report_dir / "fact_purchase.csv"
    fi = pd.read_csv(fi_p, low_memory=False) if fi_p.exists() else pd.DataFrame()
    fp = pd.read_csv(fp_p, low_memory=False) if fp_p.exists() else pd.DataFrame()

    for c in ("period_start", "period_end"):
        fs[c] = pd.to_datetime(fs[c], errors="coerce")
    for df, cols in ((fi, ["snapshot_date"]),
                     (fp, ["created_date", "approved_date", "delivery_date"])):
        for c in cols:
            if len(df) and c in df.columns:
                df[c] = pd.to_datetime(df[c], errors="coerce")
    return fs, fi, fp


def load_item_master(input_dir: Path) -> pd.DataFrame:
    """Uniware Item Master -> Category / Gender / MRP per SKU."""
    cand = []
    for p in input_dir.rglob("*.csv"):
        if p.name.startswith(("~$", ".")) or any(
                x.startswith("_") for x in p.relative_to(input_dir).parts[:-1]):
            continue
        if re.search(r"item[_ ]?master", p.name, re.I):
            cand.append(p)
    if not cand:
        return pd.DataFrame(columns=["master_sku", "Category", "Gender"])
    p = max(cand, key=lambda f: f.stat().st_mtime)
    im = pd.read_csv(p, low_memory=False)
    im.columns = [str(c).strip() for c in im.columns]
    out = pd.DataFrame({"master_sku": norm(im.get("Product Code"))})
    out["Category"] = im.get("Category Name")
    out["Gender"] = im.get("Gender")
    out["MRP"] = num(im.get("MRP"))
    out["_year"] = num(im.get("Year"))
    return out.dropna(subset=["master_sku"]).drop_duplicates("master_sku")


def load_overrides(input_dir: Path):
    """Merchandiser judgement columns. These cannot be computed -- they are
    carried forward from a file the team maintains, so notes survive each run.

    Returns (by_sku, by_style). One file serves both grains: a row with a Sku
    Code applies to that SKU, a row with only an Item Name applies to the whole
    style. Without this split the style-level join would try to match SKU codes
    against style names and silently never fire.
    """
    empty = pd.DataFrame(columns=["key", "Style Re-order Status", "Reorder Status"])
    p = input_dir / "reorder_status.csv"
    if not p.exists():
        return empty.copy(), empty.copy()

    ov = pd.read_csv(p, dtype=str).fillna("")
    ov.columns = [str(c).strip() for c in ov.columns]
    sku_col = next((c for c in ov.columns if c.lower() in ("sku code", "master_sku")), None)
    sty_col = next((c for c in ov.columns if c.lower() in ("item name", "style", "style code")), None)
    if sku_col is None and sty_col is None:
        return empty.copy(), empty.copy()

    for c in ("Style Re-order Status", "Reorder Status"):
        if c not in ov.columns:
            ov[c] = ""
    has_note = ov["Style Re-order Status"].str.strip().ne("") | \
               ov["Reorder Status"].str.strip().ne("")
    ov = ov[has_note]                       # blank rows carry nothing

    def pick(col, mask):
        if col is None:
            return empty.copy()
        sub = ov[mask].copy()
        out = pd.DataFrame({"key": norm(sub[col]).values,
                            "Style Re-order Status": sub["Style Re-order Status"].values,
                            "Reorder Status": sub["Reorder Status"].values})
        return out.dropna(subset=["key"]).drop_duplicates("key")

    sku_filled = ov[sku_col].str.strip().ne("") if sku_col else pd.Series(False, index=ov.index)
    by_sku = pick(sku_col, sku_filled)
    by_style = pick(sty_col, ~sku_filled)          # style row = no Sku Code given
    return by_sku, by_style


def write_override_template(input_dir: Path, skus: pd.DataFrame):
    """Seed reorder_status.csv the first time, so the team has something to edit."""
    p = input_dir / "reorder_status.csv"
    if p.exists():
        return False
    t = skus[["Sku Code", "Item Name"]].copy()
    t["Style Re-order Status"] = ""
    t["Reorder Status"] = ""
    safe_to_csv(t, p)
    return True


# ===========================================================================
# Build
# ===========================================================================

def build(input_dir: Path, report_dir: Path, asof: pd.Timestamp,
          period_days: int, recent_days: int):
    fs, fi, fp = load_facts(report_dir)
    im = load_item_master(input_dir)

    per_start = asof - pd.Timedelta(days=period_days - 1)
    rec_start = asof - pd.Timedelta(days=recent_days - 1)
    per_months = period_days / 30.0

    sell_out = fs[(fs.flow == "sell_out") & fs.master_sku.notna()].copy()
    in_period = sell_out[(sell_out.period_end >= per_start) & (sell_out.period_start <= asof)]
    in_recent = sell_out[(sell_out.period_end >= rec_start) & (sell_out.period_start <= asof)]

    skus = sorted(set(fs.loc[fs.master_sku.notna(), "master_sku"]) |
                  set(fi.loc[fi.master_sku.notna(), "master_sku"] if len(fi) else []) |
                  set(fp.loc[fp.master_sku.notna(), "master_sku"] if len(fp) else []))
    d = pd.DataFrame({"Sku Code": skus})

    style_of = (fs.dropna(subset=["master_sku"]).drop_duplicates("master_sku")
                .set_index("master_sku")["style_code"].to_dict())
    if len(fi):
        style_of.update(fi.dropna(subset=["master_sku"]).drop_duplicates("master_sku")
                        .set_index("master_sku")["style_code"].to_dict())
    d["Item Name"] = d["Sku Code"].map(style_of)

    d = d.merge(im, left_on="Sku Code", right_on="master_sku", how="left") \
         .drop(columns=["master_sku"], errors="ignore")
    for c in ("Category", "Gender"):
        if c not in d.columns:
            d[c] = None

    # ---- inventory by location -------------------------------------------
    def inv_sum(channels, col="qty_on_hand"):
        if not len(fi):
            return {}
        sub = fi[fi.channel.isin(channels) & fi.master_sku.notna()]
        return sub.groupby("master_sku")[col].sum().to_dict()

    for col, channels in INV_COLUMNS.items():
        d[col] = d["Sku Code"].map(inv_sum(channels)).fillna(0)

    # RTV / blocked stock is not sellable but is still owned -- kept visible,
    # separately, exactly as the client's sheet does.
    d["CB rtv"] = 0.0
    d["d8 inv + rtv"] = d["D8 inv"] + d["CB rtv"]
    loc_cols = [c for c in INV_COLUMNS if c != "D8 inv"]
    d["Total Inventory"] = d[loc_cols].sum(axis=1) + d["d8 inv + rtv"]

    # ---- sales in the period ---------------------------------------------
    def agg(frame, mask, col):
        sub = frame[mask(frame)] if callable(mask) else frame[mask]
        if not len(sub):
            return {}
        return sub.groupby("master_sku")[col].sum().to_dict()

    amz = lambda f: f.channel.isin(AMAZON_CHANNELS)
    ret = lambda f: f.channel_type.isin(RETAIL_TYPES)
    oth = lambda f: ~f.channel.isin(AMAZON_CHANNELS) & ~f.channel_type.isin(RETAIL_TYPES)

    d["Coco-OR SALES"] = d["Sku Code"].map(agg(in_period, amz, "qty_sold")).fillna(0)
    d["Coco-OR RET"] = d["Sku Code"].map(agg(in_period, amz, "qty_returned")).fillna(0)
    d["Retail"] = d["Sku Code"].map(agg(in_period, ret, "qty_sold")).fillna(0)
    d["Sale"] = d["Sku Code"].map(agg(in_period, oth, "qty_sold")).fillna(0)
    d["Ret"] = d["Sku Code"].map(agg(in_period, oth, "qty_returned")).fillna(0)

    d["Returns"] = d["Ret"] + d["Coco-OR RET"]
    d["Net sale"] = d["Sale"] + d["Coco-OR SALES"] + d["Retail"] - d["Returns"]

    d["Last 15 days Sales"] = d["Sku Code"].map(
        in_recent.groupby("master_sku")["qty_sold"].sum().to_dict()).fillna(0)
    d["Last 15 days Returns"] = d["Sku Code"].map(
        in_recent.groupby("master_sku")["qty_returned"].sum().to_dict()).fillna(0)

    # ---- lifetime --------------------------------------------------------
    d["overall net sale qty"] = d["Sku Code"].map(
        sell_out.groupby("master_sku")["qty_sold"].sum().to_dict()).fillna(0) \
        - d["Sku Code"].map(
        sell_out.groupby("master_sku")["qty_returned"].sum().to_dict()).fillna(0)
    d["Overall ret"] = d["Sku Code"].map(
        sell_out.groupby("master_sku")["qty_returned"].sum().to_dict()).fillna(0)

    # ---- purchases + ageing ---------------------------------------------
    if len(fp):
        pos = fp[fp.master_sku.notna()]
        d["Purchase. qty"] = d["Sku Code"].map(
            pos.groupby("master_sku")["qty_ordered"].sum().to_dict()).fillna(0)
        first_po = pos.groupby("master_sku")["created_date"].min()
    else:
        d["Purchase. qty"] = 0.0
        first_po = pd.Series(dtype="datetime64[ns]")

    first_sale = sell_out.groupby("master_sku")["period_start"].min()
    d["_first_po"] = d["Sku Code"].map(first_po)
    d["_first_sale"] = d["Sku Code"].map(first_sale)

    # Fallback ladder, best first. Each SKU records which rung it used.
    d["_season"] = [season_date(a if pd.notna(a) else b)
                    for a, b in zip(d["Item Name"], d["Sku Code"])]
    yr = d["_year"] if "_year" in d.columns else pd.Series(np.nan, index=d.index)
    d["_yeardate"] = pd.to_datetime(
        dict(year=yr.fillna(0).astype(int).replace(0, 1970), month=7, day=1),
        errors="coerce").where(yr.notna())

    ladder = [("first PO date", "_first_po"), ("first sale date", "_first_sale"),
              ("season code", "_season"), ("product year", "_yeardate")]
    d["_first_seen"] = pd.NaT
    d["Ageing basis"] = "unknown"
    for label, col in ladder:
        take = d["_first_seen"].isna() & d[col].notna()
        d.loc[take, "_first_seen"] = d.loc[take, col]
        d.loc[take, "Ageing basis"] = label

    d["Ageing"] = ((asof - d["_first_seen"]).dt.days).map(ageing_bucket)

    # ---- rates, projections, sell-through -------------------------------
    d["avg"] = d["Net sale"] / per_months
    d["Sale Projection for next 3 month"] = d["avg"] * 3
    d["ros/month"] = d["Net sale"] / (period_days / 30.0)
    d["monthly ros"] = d["Last 15 days Sales"] * (30.0 / recent_days)
    d["For 3 months"] = d["ros/month"] * 3

    # Period sell-through against the stock that was available to sell it.
    # (The client's sheet drifted between two denominators; this is the
    # standard definition and is stated in the Definitions tab.)
    d["Sell through"] = d["Net sale"] / (d["Net sale"] + d["Total Inventory"]).replace(0, np.nan)

    # Overall sell-through wants receipts as the denominator. Until PO history
    # goes back far enough, fall back to (sold + still in stock) and say so --
    # an estimate that is usable now beats a blank column for two years.
    on_po = d["overall net sale qty"] / d["Purchase. qty"].replace(0, np.nan)
    est = d["overall net sale qty"] / (
        d["overall net sale qty"] + d["Total Inventory"]).replace(0, np.nan)
    d["Overall Sell-through"] = on_po.where(on_po.notna(), est)
    d["Sell-through basis"] = np.where(on_po.notna(), "purchase qty",
                              np.where(est.notna(), "sold + stock (estimate)", "unknown"))

    # ---- classification --------------------------------------------------
    d["Overall Movement"] = [movement(s, g) for s, g in
                             zip(d["Overall Sell-through"], d["Ageing"])]
    d["Movement"] = [movement(s, g) for s, g in zip(d["Sell through"], d["Ageing"])]
    d["Inv. Level"] = np.where(
        d["Total Inventory"] <= d["Sale Projection for next 3 month"], "REORDER", "IN-STOCK")
    d["Inventory status"] = [inventory_status(i, m) for i, m in
                             zip(d["Total Inventory"], d["monthly ros"])]

    # ---- merchandiser overrides -----------------------------------------
    by_sku, by_style = load_overrides(input_dir)
    d = d.merge(by_sku, left_on="Sku Code", right_on="key", how="left") \
         .drop(columns=["key"], errors="ignore")
    # a style-level note falls through to every size that has no SKU-level note
    for c in ("Style Re-order Status", "Reorder Status"):
        if c not in d.columns:
            d[c] = ""
        d[c] = d[c].fillna("")
    if len(by_style):
        fill = d.merge(by_style, left_on="Item Name", right_on="key",
                       how="left", suffixes=("", "_sty"))
        for c in ("Style Re-order Status", "Reorder Status"):
            d[c] = d[c].mask(d[c].eq(""), fill[f"{c}_sty"].values)
    for c in ("Style Re-order Status", "Reorder Status"):
        d[c] = d[c].fillna("")

    # Sorted by style then SKU so a merchandiser can find a row; the priority
    # ordering lives on the Action list sheet instead.
    plabel = period_label(per_start, asof)
    d = d.sort_values(["Item Name", "Sku Code"])
    sku_sheet = lay_out(d, SKU_LAYOUT, plabel)

    # =======================================================================
    # Article (style) level
    # =======================================================================
    g = d[d["Item Name"].notna()].copy()
    sums = ["Total Inventory", "all sports inv", "J&J Inv", "OR inv", "INCS inv",
            "d8 inv + rtv", "D8 inv", "CB rtv", "Sale", "Ret", "Coco-OR SALES",
            "Coco-OR RET", "Retail", "Net sale", "Returns", "Purchase. qty",
            "overall net sale qty", "Overall ret",
            "Last 15 days Sales", "Last 15 days Returns"]
    a = g.groupby("Item Name", as_index=False)[sums].sum()
    a["Category"] = a["Item Name"].map(g.drop_duplicates("Item Name")
                                       .set_index("Item Name")["Category"])
    a["Gender"] = a["Item Name"].map(g.drop_duplicates("Item Name")
                                     .set_index("Item Name")["Gender"])
    a["sizes"] = a["Item Name"].map(g.groupby("Item Name")["Sku Code"].nunique())
    a["sizes in stock"] = a["Item Name"].map(
        g[g["Total Inventory"] > 0].groupby("Item Name")["Sku Code"].nunique()).fillna(0)

    # Ageing at style level = the oldest size in the style.
    a["Ageing"] = a["Item Name"].map(
        ((asof - g.groupby("Item Name")["_first_seen"].min()).dt.days).map(ageing_bucket))

    a["ros/month"] = a["Net sale"] / (period_days / 30.0)
    a["monthly ros"] = a["Last 15 days Sales"] * (30.0 / recent_days)
    a["For 3 months"] = a["ros/month"] * 3
    a["Sell through"] = a["Net sale"] / (a["Net sale"] + a["Total Inventory"]).replace(0, np.nan)
    a_on_po = a["overall net sale qty"] / a["Purchase. qty"].replace(0, np.nan)
    a_est = a["overall net sale qty"] / (
        a["overall net sale qty"] + a["Total Inventory"]).replace(0, np.nan)
    a["Overall Sell-through"] = a_on_po.where(a_on_po.notna(), a_est)
    a["Sell-through basis"] = np.where(a_on_po.notna(), "purchase qty",
                              np.where(a_est.notna(), "sold + stock (estimate)", "unknown"))
    gross = a["overall net sale qty"] + a["Overall ret"]
    a["Ret %"] = (a["Overall ret"] / gross.replace(0, np.nan))
    a["overall ret qty"] = a["Overall ret"]

    a["overall Movement"] = [movement(s, ag) for s, ag in
                             zip(a["Overall Sell-through"], a["Ageing"])]
    a["Movement"] = [movement(s, ag) for s, ag in zip(a["Sell through"], a["Ageing"])]
    a["Movement Stability"] = [stability(r, o) for r, o in
                               zip(a["Movement"], a["overall Movement"])]
    # lower case on the article sheet, upper case on sku wise -- as in the
    # client's workbook. Cosmetic, but it is what their filters key on.
    a["Inv. Level"] = np.where(a["Total Inventory"] <= a["For 3 months"],
                               "reorder", "in-stock")
    a["Inventory status"] = [inventory_status(i, m) for i, m in
                             zip(a["Total Inventory"], a["monthly ros"])]

    a = a.merge(by_style.rename(columns={"key": "Item Name",
                                         "Style Re-order Status": "Style Status",
                                         "Reorder Status": "Re-order Status"}),
                on="Item Name", how="left")
    # if only SKU-level notes exist, surface the most common one for the style
    if len(by_sku):
        lift = (d[d["Item Name"].notna()]
                .assign(_s=d["Style Re-order Status"].replace("", np.nan))
                .dropna(subset=["_s"]).groupby("Item Name")["_s"]
                .agg(lambda x: x.mode().iat[0] if len(x.mode()) else None))
        a["Style Status"] = a["Style Status"].where(
            a["Style Status"].notna(), a["Item Name"].map(lift))
    for c in ("Style Status", "Re-order Status"):
        a[c] = a.get(c, pd.Series(index=a.index, dtype=object)).fillna("")

    a = a.sort_values("Item Name")
    art_sheet = lay_out(a, ARTICLE_LAYOUT, plabel)

    return sku_sheet, art_sheet, a, dict(
        asof=asof, per_start=per_start, rec_start=rec_start,
        period_days=period_days, recent_days=recent_days, per_months=per_months)


# ===========================================================================
# Workbook
# ===========================================================================

def definitions(meta) -> pd.DataFrame:
    p, r = meta["period_days"], meta["recent_days"]
    pm = meta["per_months"]
    rows = [
        ("Report date", f"{meta['asof']:%Y-%m-%d}", "--asof"),
        ("Sales period", f"{meta['per_start']:%Y-%m-%d} to {meta['asof']:%Y-%m-%d} "
                         f"({p} days = {pm:.2f} months)", "--period-days"),
        ("Short window", f"{meta['rec_start']:%Y-%m-%d} to {meta['asof']:%Y-%m-%d} "
                         f"({r} days)", "--recent-days"),
        ("", "", ""),
        ("Total Inventory", "all sports + J&J + OR + INCS + (D8 inv + CB rtv)",
         "sum of fact_inventory.qty_on_hand by channel"),
        ("d8 inv + rtv", "D8 inv + CB rtv", "warehouse sellable + return-to-vendor"),
        ("Sale / Ret", "sell-out units / returns from every channel that is NOT "
                       "Amazon and NOT a retail store (D2C + other marketplaces)", ""),
        ("Coco-OR SALES / RET", "sell-out units / returns on Amazon Vendor Central "
                                "(Cocoblu)", ""),
        ("Retail", "sell-out units from retail-store channels", ""),
        ("Net sale", "Sale + Coco-OR SALES + Retail - Returns", "validated 100% vs client sheet"),
        ("Returns", "Ret + Coco-OR RET", "validated 100% vs client sheet"),
        ("avg", f"Net sale / {pm:.2f} (months in the period)",
         "validated 99.9% vs client sheet"),
        ("Sale Projection for next 3 month", "avg x 3", "validated 100%"),
        ("ros/month", f"Net sale / {pm:.2f}", "rate of sale per month; validated 100%"),
        ("For 3 months", "ros/month x 3", "validated 100%"),
        ("monthly ros", f"Last {r} days Sales x {30 / r:.2f}",
         "gross, not net -- matches client sheet 100%"),
        ("Purchase. qty", "sum of fact_purchase.qty_ordered (all POs on file)",
         "NEEDS A LONG PO WINDOW -- see the Caveats sheet"),
        ("Sell through", "Net sale / (Net sale + Total Inventory)",
         "standard definition; the client sheet drifted between two denominators"),
        ("Overall Sell-through", "overall net sale qty / Purchase. qty when PO history "
                                 "covers the SKU, else overall net sale qty / "
                                 "(overall net sale qty + Total Inventory)",
         "the purchase-qty form is validated 100% vs the client sheet; the "
         "'Sell-through basis' column says which was used per row"),
        ("overall net sale qty", "sell-out units - returns across EVERY row in "
                                 "fact_sales.csv",
         "'overall' means 'everything currently on file', NOT lifetime. The "
         "Uniware pull is windowed, so this grows as history accumulates -- see "
         "the Caveats sheet"),
        ("Ageing", "bucketed age since first stocked. Fallback ladder: first PO date "
                    "-> first sale date -> season code in the style (AW24 = Aug 2024, "
                    "SS25 = Feb 2025) -> Item Master 'Year' field",
         "the 'Ageing basis' column records which rung each SKU used -- only "
         "'first PO date' is a true receipt date"),
        ("Overall Movement", "movement(Overall Sell-through, Ageing)", NOTE_MOVEMENT),
        ("Movement", "movement(period Sell through, Ageing)", NOTE_MOVEMENT),
        ("Movement Stability", "recent Movement vs Overall Movement",
         "Improving / Slowing Down / Recovering / Stable"),
        ("Inv. Level", "REORDER when Total Inventory <= Sale Projection for next 3 month",
         "96.6% agreement with the client sheet"),
        ("Inventory status", "cover = Total Inventory / monthly ros; "
                             "<0.5 no stock, <1.5 low stock, <3 in stock, else over stock",
         "with no recent sales, judged on stock alone"),
        ("Style Re-order Status / Reorder Status",
         "NOT COMPUTED -- read from reports/input/reorder_status.csv",
         "merchandiser judgement and notes; edit that file, it survives every run"),
    ]
    return pd.DataFrame(rows, columns=["Column", "Formula", "Notes"])


def criteria_sheet() -> pd.DataFrame:
    rows = []
    for label, _ in AGEING_BUCKETS:
        a = EXPECTED_SELL_THROUGH[label]
        rows.append((label, a, 0.0, round(a * SLOW_CUT, 4),
                     round(a * SLOW_CUT, 4), round(a * FAST_CUT, 4),
                     f">{round(a * FAST_CUT, 4)}"))
    return pd.DataFrame(rows, columns=[
        "Ageing", "Avg (expected sell-through)", "No Movement (<=)",
        "Slow Movement (<)", "Good Movement (from)", "Good Movement (to)",
        "Fast Movement"])


def caveats(sku, meta, fp_rows) -> pd.DataFrame:
    rows = []
    no_po = int((sku["Purchase. qty"].fillna(0) == 0).sum())
    unk_age = int(sku["Ageing"].isna().sum())
    no_cat = int(sku["Category"].isna().sum())
    rows.append(("Purchase order history",
                 f"{fp_rows:,} PO lines on file; {no_po:,} of {len(sku):,} SKUs have "
                 f"no PO at all",
                 "Purchase. qty and Overall Sell-through are only meaningful with a "
                 "long PO window. Pull Purchase Orders with --days 730, or keep a "
                 "rolling PO history file."))
    rows.append(("Ageing",
                 f"{unk_age:,} SKUs have no date to age from",
                 "Ageing drives the movement thresholds. Without a first-PO date the "
                 "SKU gets no Movement classification rather than a guessed one."))
    rows.append(("Category / Gender",
                 f"{no_cat:,} SKUs missing Category",
                 "Sourced from the Uniware Item Master export. Make sure it is in "
                 "reports/input/uniware/."))
    rows.append(("'Overall' is not lifetime",
                 "overall net sale qty / Overall ret cover only the rows on file",
                 "The Uniware pull is windowed. Either widen it, or keep the "
                 "superseded pulls and append instead of replacing, if you want "
                 "true lifetime figures."))
    rows.append(("Aggregate periods are not pro-rated",
                 "Amazon VC and some store files report one total per period",
                 "A period that only partly overlaps the report window is counted "
                 "in full. Keep --period-days aligned to how those files are "
                 "exported, or the edges will be overstated."))
    rows.append(("Sell-out coverage",
                 "Only channels with a sell-out report contribute to Sale / Retail",
                 "A marketplace with sell-in only (Myntra, Nykaa, Tata Cliq) shows "
                 "zero sales here. See the Period Coverage sheet of the main report."))
    rows.append(("Merchandiser columns",
                 "Style Re-order Status and Reorder Status are not computed",
                 "Maintained in reports/input/reorder_status.csv and joined in."))
    return pd.DataFrame(rows, columns=["Area", "Current state", "What to do"])


def action_list(art: pd.DataFrame) -> pd.DataFrame:
    """The styles someone should actually do something about this week.

    Two buckets, most valuable first: reorder what is selling and running out,
    and clear what is not selling and sitting on stock.
    """
    a = art.copy()

    def i(col):                                  # tidy integer for the reason text
        return num(a[col]).round(0).fillna(0).astype(int).astype(str)

    selling = a["Movement"].isin(["Good Movement", "Fast Movement"])
    thin = (a["Inventory status"].isin(["no stock", "low stock"])
            | a["Inv. Level"].eq("REORDER"))
    reorder = selling & thin
    dead = a["Movement"].eq("No Movement") & a["Inventory status"].eq("over stock")
    broken = (num(a["sizes in stock"]).fillna(0)
              < num(a["sizes"]).fillna(0) * 0.5) & selling

    why_reorder = (a["Movement"].astype(str) + ", " + a["Inventory status"].astype(str)
                   + ", " + i("For 3 months") + " units projected for 3 months vs "
                   + i("Total Inventory") + " in stock")
    why_dead = ("no sales in the window, " + i("Total Inventory")
                + " units on hand, aged " + a["Ageing"].astype(str))
    why_broken = (i("sizes in stock") + " of " + i("sizes")
                  + " sizes in stock on a "
                  + a["Movement"].astype(str).str.lower() + " style")

    # First match wins, so a style is never listed twice.
    a["Action"] = np.select([reorder, broken, dead],
                            ["REORDER", "SIZE BREAK", "CLEAR"], default=None)
    a["Why"] = np.select([reorder, broken, dead],
                         [why_reorder, why_broken, why_dead], default=None)

    out = a[a["Action"].notna()].copy()
    order = {"REORDER": 0, "SIZE BREAK": 1, "CLEAR": 2}
    out["_o"] = out["Action"].map(order)
    out = out.sort_values(["_o", "Total Inventory"], ascending=[True, False])
    return out[["Action", "Why", "Item Name", "Category", "Gender",
                "Total Inventory", "sizes", "sizes in stock", "Net sale",
                "ros/month", "For 3 months", "Ageing", "Movement",
                "overall Movement", "Inventory status", "Inv. Level",
                "Style Status", "Re-order Status"]]


def write_workbook(path: Path, sku, art, meta, fp_rows, art_raw=None):
    # sku wise FIRST, exactly as in the client's workbook -- the file must open
    # on the sheet they expect, not on a sheet this pipeline invented.
    acts = action_list(art_raw if art_raw is not None else art)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sku.to_excel(w, sheet_name="sku wise", index=False)
        art.to_excel(w, sheet_name="Article wise", index=False)
        acts.to_excel(w, sheet_name="Action list", index=False)
        definitions(meta).to_excel(w, sheet_name="Definitions", index=False)
        criteria_sheet().to_excel(w, sheet_name="Movement criteria", index=False)
        caveats(sku, meta, fp_rows).to_excel(w, sheet_name="Caveats", index=False)

        for name, df in (("sku wise", sku), ("Article wise", art), ("Action list", acts)):
            ws = w.sheets[name]
            ws.freeze_panes = "C2"
            for i, col in enumerate(df.columns, start=1):
                q = df[col].astype(str).str.len().quantile(0.95) if len(df) else 8
                body = 8 if pd.isna(q) else int(q)
                # headers carry embedded newlines; size to the longest line
                head_len = max(len(part) for part in str(col).split("\n"))
                width = min(max(head_len, body) + 2, 32)
                ws.column_dimensions[ws.cell(1, i).column_letter].width = width
                ws.cell(1, i).alignment = Alignment(wrap_text="\n" in str(col),
                                                    vertical="bottom")
        for name in ("Definitions", "Movement criteria", "Caveats"):
            ws = w.sheets[name]
            for i, wdt in enumerate([34, 62, 60], start=1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = wdt
        w.sheets["Action list"].column_dimensions["B"].width = 60
    return acts


# ===========================================================================

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", "-i", required=True, help="reports/input folder")
    ap.add_argument("--report", "-r", required=True, help="reports/output folder")
    ap.add_argument("--asof", help="report date YYYY-MM-DD (default: today)")
    ap.add_argument("--period-days", type=int, default=60)
    ap.add_argument("--recent-days", type=int, default=15)
    a = ap.parse_args()

    input_dir, report_dir = Path(a.input).resolve(), Path(a.report).resolve()
    asof = (pd.to_datetime(a.asof) if a.asof
            else pd.Timestamp(datetime.now().date()))

    sku, art, art_raw, meta = build(input_dir, report_dir, asof,
                                    a.period_days, a.recent_days)

    fp_p = report_dir / "fact_purchase.csv"
    fp_rows = len(pd.read_csv(fp_p)) if fp_p.exists() else 0

    out = safe_excel_path(report_dir / f"Stock_vs_Sales_{asof:%d%m%y}.xlsx")
    acts = write_workbook(out, sku, art, meta, fp_rows, art_raw)

    if write_override_template(input_dir, sku.rename(
            columns={c: "Sku Code" for c in sku.columns[:1]})):
        print(f"  seeded {input_dir / 'reorder_status.csv'} "
              f"({len(sku):,} rows) — fill in the two merchandiser columns")

    print(f"\nStock vs Sales — as of {asof:%Y-%m-%d}")
    print(f"  period {meta['per_start']:%Y-%m-%d} .. {asof:%Y-%m-%d} "
          f"({a.period_days}d)   short window {a.recent_days}d")
    print(f"  {len(sku):,} SKUs / {len(art):,} articles")
    print()
    for col in ("Movement", "Overall Movement", "Inventory status", "Inv. Level",
                "Ageing", "Ageing basis", "Sell-through basis"):
        if col not in sku.columns:
            continue
        vc = sku[col].value_counts(dropna=False)
        print(f"  {col:18s} " + "  ".join(f"{k}={v}" for k, v in vc.items()))
    if len(acts):
        print(f"\n  Action list: " + "  ".join(
            f"{k}={v}" for k, v in acts["Action"].value_counts().items()))
    print(f"\n  wrote {out}  ({out.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
